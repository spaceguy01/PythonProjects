""" Working with Excel Files using openpyxl """
import openpyxl

"""WORKING WITH ALREADY EXISTING EXCEL FILE """

""" Open excel file as workbook in same directory """
worksheet = openpyxl.load_workbook('example.xlsx')


""" Getting sheetnames of workbook and set sheets to variable"""
print(worksheet.sheetnames)
sheet1 = worksheet['Sheet1']
sheet2 = worksheet['Sheet2']

""" Setting Cells as variables in Two ways"""
a1 = sheet1['A1']
A1 = sheet1.cell(row=1, column=1)

b2 = sheet1['B2']
B2 = sheet1.cell(row=2, column=2)

""" Getting Values of Cells """
a1.value
A1.value

""" Using for loop to get values of cells in B column"""
for i in range(1,9):
    print (i, sheet1.cell(row=i, column=2).value)


"""------------------------------------------------------"""
""" CREATING NEW EXCEL FILE """

workbook = openpyx.Workbook()

""" Create new Excel Sheet and set it to a variable """
sheet1 = workbook.create_sheet()

""" Set new title to newly created Sheet1 """
sheet1.title = "NewSheet"

""" Setting values into Cell in new sheet """
sheet1['A1'].value = "THIS IS NEW VALUE"

""" Create New Sheet and Insert it into specific index """
workbook.create_sheet(index=1, title="Another New Sheet")


""" Save the Modified Excel File """
workbook.save('filename.xlsx')
